from pathlib import Path
from typing import Any
from typing import List
from typing import Optional
from typing import Tuple
from typing import Union
from unittest import TestLoader
from unittest import TestSuite
from xlform.engine.base import Book
from xlform.engine.base import Cell
from xlform.engine.base import Engine
from xlform.engine.base import Range
from xlform.engine.base import Sheet
from xlform.exception import XlFormArgumentException
from xlform.exception import XlFormNotImplementedException
import openpyxl  # type: ignore
import tempfile
import unittest


class EngineTestCase(unittest.TestCase):
    def setUp(self) -> None:
        class FakeEngine(Engine):
            def new_book(self) -> Book:
                pass

            def open_book(self, path: Path) -> Book:
                pass

        self._engine: Engine = FakeEngine()

    def _set_ws_values(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        rows: List[List[Union[float, int, str]]],
        column_left: int = 1,
        row_top: int = 1,
    ) -> None:
        for y, row in enumerate(rows, start=row_top):
            for x, value in enumerate(row, start=column_left):
                ws.cell(row=y, column=x, value=value)

    def _get_book_path(
        self,
        prefix: Optional[str] = None,
        rows: Optional[List[List[Union[float, int, str]]]] = None,
    ) -> Path:
        wb = openpyxl.Workbook()
        assert wb.sheetnames == ["Sheet"]

        # Modify the cell to avoid no output of the file
        ws = wb.active
        if rows is None:
            rows = [[0]]
        self._set_ws_values(ws, rows)

        tmp_dir_path = Path(tempfile.mkdtemp(prefix=prefix))
        path = tmp_dir_path / "book.xlsx"
        wb.save(str(path))
        wb.close()

        return path

    def _get_a1_zero_book_path(self, prefix: Optional[str] = None) -> Path:
        return self._get_book_path(prefix=prefix)

    def test_engine_init(self) -> None:
        self.assertIsInstance(self._engine, Engine)

    def test_engine_new_book(self) -> None:
        book: Book = self._engine.new_book()

        self.assertIsInstance(book, Book)

    def test_engine_new_book__new_book_has_only_one_sheet(self) -> None:
        book: Book = self._engine.new_book()
        self.assertIsInstance(book, Book)
        sheets: List[Sheet] = book.get_sheets()

        self.assertEqual(len(sheets), 1)
        self.assertIsInstance(sheets[0], Sheet)

    def test_engine_new_book__sheet_name_is_sheet1(self) -> None:
        book: Book = self._engine.new_book()
        self.assertIsInstance(book, Book)
        sheets: List[Sheet] = book.get_sheets()

        # In the case of openpyxl, the name of the sheet is 'Sheet'.
        # In the case of Microsoft Excel, the name of the sheet is 'Sheet1'.
        # The Engine uses 'Sheet1' as the name of the sheet.
        self.assertEqual(sheets[0].get_name(), "Sheet1")

    def test_engine_open_book(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_engine_open_book")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()

        self.assertEqual(len(sheets), 1)
        self.assertIsInstance(sheets[0], Sheet)
        self.assertEqual(sheets[0].get_name(), "Sheet")

    def test_book_save__file_exists(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_book_save")

        book: Book = self._engine.open_book(path)
        path2 = Path(tempfile.mkdtemp()) / "book.xlsx"
        book.save(path2)

        self.assertTrue(path2.exists())

    def test_book_save__a1_is_zero(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_book_save")
        book: Book = self._engine.open_book(path)
        path2 = Path(tempfile.mkdtemp()) / "book.xlsx"
        book.save(path2)

        wb = openpyxl.load_workbook(str(path2))
        ws = wb.active
        self.assertEqual(ws["A1"].value, 0)

    def test_book_close(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_book_close")

        book: Book = self._engine.open_book(path)
        book.close()

    def test_book_iter_sheets(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_book_iter_sheets")

        wb = openpyxl.load_workbook(str(path))
        wb.create_sheet("Sheet2")
        wb.save(str(path))
        wb.close()

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = list(book.iter_sheets())

        self.assertEqual(len(sheets), 2)
        self.assertIsInstance(sheets[0], Sheet)
        self.assertIsInstance(sheets[1], Sheet)
        self.assertEqual(sheets[1].get_name(), "Sheet2")

    def test_book_get_sheets(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_book_get_sheets")

        wb = openpyxl.load_workbook(str(path))
        wb.create_sheet("Sheet2")
        wb.save(str(path))
        wb.close()

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()

        self.assertIsInstance(sheets, list)
        self.assertEqual(len(sheets), 2)
        self.assertIsInstance(sheets[0], Sheet)
        self.assertIsInstance(sheets[1], Sheet)
        self.assertEqual(sheets[1].get_name(), "Sheet2")

    def test_book_add_sheet(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_book_add_sheet")

        book: Book = self._engine.open_book(path)
        book.add_sheet("Sheet2")
        path2 = Path(tempfile.mkdtemp()) / "book.xlsx"
        book.save(path2)

        wb = openpyxl.load_workbook(str(path2))
        self.assertEqual(len(wb.sheetnames), 2)

    def test_sheet_get_name(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_sheet_get_name")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()

        self.assertEqual(len(sheets), 1)
        self.assertEqual(sheets[0].get_name(), "Sheet")

    def test_sheet_get_range__simple(self) -> None:
        path = self._get_book_path(
            rows=[[11, 12], [21, 22], [31, 32]], prefix="test_sheet_get_range"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        r: Range = sheet.get_range("A1:B3")

        self.assertIsInstance(r, Range)
        self.assertEqual(r.get_rows_count(), 3)
        self.assertEqual(r.get_columns_count(), 2)

    def test_sheet_get_range__single_column(self) -> None:
        path = self._get_book_path(
            rows=[[11, 12], [21, 22], [31, 32]], prefix="test_sheet_get_range"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        r: Range = sheet.get_range("A:A")

        self.assertIsInstance(r, Range)
        self.assertEqual(r.get_rows_count(), 3)
        self.assertEqual(r.get_columns_count(), 1)

    def test_sheet_get_range__single_row(self) -> None:
        path = self._get_book_path(
            rows=[[11, 12], [21, 22], [31, 32]], prefix="test_sheet_get_range"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        r: Range = sheet.get_range("1:1")

        self.assertIsInstance(r, Range)
        self.assertEqual(r.get_rows_count(), 1)
        self.assertEqual(r.get_columns_count(), 2)

    def test_sheet_get_range__single_cell(self) -> None:
        path = self._get_book_path(
            rows=[[11, 12], [21, 22], [31, 32]], prefix="test_sheet_get_range"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        r: Range = sheet.get_range("B1")

        self.assertIsInstance(r, Range)
        self.assertEqual(r.get_rows_count(), 1)
        self.assertEqual(r.get_columns_count(), 1)

    def test_sheet_get_cell(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_sheet_get_cell")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)

        self.assertIsInstance(c, Cell)
        self.assertEqual(c.get_value(), 0)

    def test_sheet_protect(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_sheet_protect")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet = sheets[0]
        sheet.protect()
        path2 = Path(tempfile.mkdtemp()) / "book.xlsx"
        book.save(path2)

        self.assertTrue(path2.exists())
        wb = openpyxl.load_workbook(str(path2))
        ws = wb.active
        self.assertTrue(ws.protection.sheet)

    def test_sheet_unprotect(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_sheet_unprotect")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet = sheets[0]
        sheet.unprotect()
        path2 = Path(tempfile.mkdtemp()) / "book.xlsx"
        book.save(path2)

        self.assertTrue(path2.exists())
        wb = openpyxl.load_workbook(str(path2))
        ws = wb.active
        self.assertFalse(ws.protection.sheet)

    def test_range_get_rows_count(self) -> None:
        path = self._get_book_path(
            rows=[[11, 12], [21, 22], [31, 32]], prefix="test_sheet_get_range"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        r: Range = sheet.get_range("A1:C2")

        self.assertEqual(r.get_rows_count(), 2)

    def test_range_get_columns_count(self) -> None:
        path = self._get_book_path(
            rows=[[11, 12], [21, 22], [31, 32]], prefix="test_sheet_get_range"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        r: Range = sheet.get_range("A1:C2")

        self.assertEqual(r.get_columns_count(), 3)

    def test_range_get_cell__origin(self) -> None:
        path = self._get_book_path(
            rows=[[11, 12], [21, 22], [31, 32]], prefix="test_range_get_cell"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        r: Range = sheet.get_range("A1:C2")
        c: Cell = r.get_cell(1, 1)

        self.assertIsInstance(c, Cell)
        self.assertEqual(c.get_value(), 11)

    def test_range_get_cell__offset(self) -> None:
        path = self._get_book_path(
            rows=[[11, 12, 13], [21, 22, 23], [31, 32, 33]],
            prefix="test_range_get_cell",
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        r: Range = sheet.get_range("B2:C3")
        c: Cell = r.get_cell(1, 1)

        self.assertIsInstance(c, Cell)
        self.assertEqual(c.get_value(), 22)

    def test_range_get_cell__out_of_row_range(self) -> None:
        path = self._get_book_path(
            rows=[[11, 12, 13], [21, 22, 23], [31, 32, 33]],
            prefix="test_range_get_cell",
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        r: Range = sheet.get_range("B2:C3")
        r.get_cell(2, 1)
        with self.assertRaises(XlFormArgumentException):
            r.get_cell(3, 1)

    def test_cell_get_formula__simple_formula(self) -> None:
        path = self._get_book_path(
            rows=[["=1+1"]], prefix="test_cell_get_formula"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        formula = c.get_formula()

        self.assertEqual(formula, "=1+1")
        self.assertIsInstance(formula, str)

    def test_cell_get_formula__type_int(self) -> None:
        path = self._get_book_path(rows=[[1]], prefix="test_cell_get_formula")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        formula = c.get_formula()

        self.assertEqual(formula, 1)
        self.assertIsInstance(formula, int)

    def test_cell_get_formula__type_float(self) -> None:
        path = self._get_book_path(
            rows=[[1.5]], prefix="test_cell_get_formula"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        formula = c.get_formula()

        self.assertEqual(formula, 1.5)
        self.assertIsInstance(formula, float)

    def test_cell_get_formula__type_str(self) -> None:
        path = self._get_book_path(
            rows=[["a"]], prefix="test_cell_get_formula"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        formula = c.get_formula()

        self.assertEqual(formula, "a")
        self.assertIsInstance(formula, str)

    def test_cell_get_value__simple_formula(self) -> None:
        path = self._get_book_path(
            rows=[["=1+1"]], prefix="test_cell_get_value"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        with self.assertRaises(XlFormNotImplementedException):
            c.get_value()

    def test_cell_get_value__type_int(self) -> None:
        path = self._get_book_path(rows=[[1]], prefix="test_cell_get_value")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        value = c.get_value()

        self.assertEqual(value, 1)
        self.assertIsInstance(value, int)

    def test_cell_get_value__type_float(self) -> None:
        path = self._get_book_path(rows=[[1.5]], prefix="test_cell_get_value")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        value = c.get_value()

        self.assertEqual(value, 1.5)
        self.assertIsInstance(value, float)

    def test_cell_get_value__type_str(self) -> None:
        path = self._get_book_path(rows=[["a"]], prefix="test_range_get_value")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        value = c.get_value()

        self.assertEqual(value, "a")
        self.assertIsInstance(value, str)

    def test_cell_get_value__type_datetime(self) -> None:
        self.skipTest("Test not implemented.")

    def test_cell_get_number_format(self) -> None:
        path = self._get_a1_zero_book_path(
            prefix="test_cell_get_number_format"
        )

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        value = c.get_number_format()

        self.assertEqual(value, "General")
        self.assertIsInstance(value, str)

    def test_cell_get_text(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_cell_get_text")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        text = c.get_text()

        self.assertEqual(text, "0")
        self.assertIsInstance(text, str)

    def test_cell_get_row(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_cell_get_row")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(2, 3)

        self.assertEqual(c.get_row(), 2)

    def test_cell_get_column(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_cell_get_column")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(2, 3)

        self.assertEqual(c.get_column(), 3)

    def test_cell_get_address(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_cell_get_address")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        address = c.get_address()

        self.assertEqual(address, "$A$1")

    def test_cell_get_address__relative(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_cell_get_address")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet: Sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        address = c.get_address(column_absolute=False, row_absolute=False)

        self.assertEqual(address, "A1")

    def test_cell_set_value(self) -> None:
        path = self._get_a1_zero_book_path(prefix="test_cell_set_value")

        book: Book = self._engine.open_book(path)
        sheets: List[Sheet] = book.get_sheets()
        sheet = sheets[0]
        c: Cell = sheet.get_cell(1, 1)
        c.set_value(1)
        path2 = Path(tempfile.mkdtemp()) / "book.xlsx"
        book.save(path2)

        self.assertEqual(c.get_value(), 1)

    def tearDown(self) -> None:
        pass


def load_tests(loader: TestLoader, test_cases: Tuple[Any]) -> TestSuite:
    suite = unittest.TestSuite()
    for test_class in test_cases:
        tests = loader.loadTestsFromTestCase(test_class)
        suite.addTests(tests)
    return suite
