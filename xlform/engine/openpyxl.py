from pathlib import Path
from typing import Any
from typing import cast
from typing import Iterator
from typing import Tuple
from xlform.engine.base import Book
from xlform.engine.base import Cell
from xlform.engine.base import CellValue
from xlform.engine.base import Engine
from xlform.engine.base import Range
from xlform.engine.base import safe_cast_cell_value
from xlform.engine.base import Sheet
from xlform.exception import XlFormArgumentException
from xlform.exception import XlFormInternalException
from xlform.exception import XlFormNotImplementedException
import openpyxl  # type: ignore


class CellOpenpyxl(Cell):
    def __init__(self, cell: openpyxl.cell.cell.Cell):
        self._cell = cell

    def get_formula(self) -> CellValue:
        return safe_cast_cell_value(self._cell.value)

    def get_value(self) -> CellValue:
        value = safe_cast_cell_value(self._cell.value)
        if isinstance(value, str) and value[0] == "=":
            raise XlFormNotImplementedException()
        return value

    def get_number_format(self) -> str:
        number_format = self._cell.number_format
        if not isinstance(number_format, str):
            raise XlFormInternalException()
        return number_format

    def get_text(self) -> str:
        raise XlFormNotImplementedException()

    def get_address(
        self, column_absolute: bool = True, row_absolute: bool = True
    ) -> str:
        return "%s%s%s%d" % (
            "$" if column_absolute else "",
            self._cell.column_letter,
            "$" if row_absolute else "",
            self._cell.row,
        )

    def set_value(self, value: CellValue) -> None:
        self._cell.value = value


class RangeOpenpyxl(Range):
    def __init__(self, r: Tuple[Tuple[openpyxl.cell.cell.Cell]]):
        if (
            (not isinstance(r, tuple))
            or (not isinstance(r[0], tuple))
            or (not isinstance(r[0][0], openpyxl.cell.cell.Cell))
        ):
            raise XlFormArgumentException()
        self._range = r
        self._column_offset = r[0][0].column - 1
        self._row_offset = r[0][0].row - 1

    def get_cell(self, row: int, column: int) -> Cell:
        if (
            row < 1
            or self.get_rows_count() < row
            or column < 1
            or self.get_columns_count() < column
        ):
            raise XlFormArgumentException()
        return CellOpenpyxl(self._range[row - 1][column - 1])

    def get_columns_count(self) -> int:
        return len(self._range[0])

    def get_rows_count(self) -> int:
        return len(self._range)


class SheetOpenpyxl(Sheet):
    def __init__(self, sheet: Any) -> None:
        self._sheet = sheet

    def get_name(self) -> str:
        return cast(str, self._sheet.title)

    def get_cell(self, row: int, column: int) -> Cell:
        if row < 1 or column < 1:
            raise XlFormArgumentException()
        return CellOpenpyxl(self._sheet.cell(row=row, column=column))

    def get_range(self, arg: str) -> Range:
        r = self._sheet[arg]
        if isinstance(r, openpyxl.cell.cell.Cell):
            return RangeOpenpyxl(((r,),))  # 'A1'
        if isinstance(r, tuple):
            if len(r) == 0 or (len(r) > 0 and isinstance(r[0], tuple)):
                r2 = cast(Tuple[Tuple[Any]], r)
                return RangeOpenpyxl(r2)  # 'A1:A2', 'A:B' or '1:2'
            if isinstance(r[0], openpyxl.cell.cell.Cell) and (
                len(r) == 1 or (len(r) > 1 and r[0].column != r[1].column)
            ):
                r3 = cast(Tuple[Any], r)
                return RangeOpenpyxl((r3,))  # '1:1'
            r4 = cast(Tuple[Any], r)
            r5 = cast(Tuple[Tuple[Any]], tuple(map(lambda c: (c,), r4)))
            return RangeOpenpyxl(r5)  # 'A:A'
        raise XlFormInternalException()

    def protect(self) -> None:
        if False:
            self._sheet.protection.enable()
        raise XlFormNotImplementedException()

    def unprotect(self) -> None:
        if False:
            self._sheet.protection.disable()
        raise XlFormNotImplementedException()


class BookOpenpyxl(Book):
    def __init__(self, book: openpyxl.workbook.workbook.Workbook) -> None:
        self._book = book

    def save(self, path: Path) -> None:
        self._book.save(str(path))
        return None

    def close(self) -> None:
        self._book.close()

    def iter_sheets(self) -> Iterator[Sheet]:
        for sheet in self._book:
            yield SheetOpenpyxl(sheet)

    def add_sheet(self, name: str) -> None:
        self._book.create_sheet(name)
        return None


class EngineOpenpyxl(Engine):
    def __init__(self) -> None:
        pass

    def new_book(self) -> Book:
        wb = openpyxl.Workbook()
        assert wb.sheetnames == ["Sheet"]
        wb["Sheet"].title = "Sheet1"
        return BookOpenpyxl(wb)

    def open_book(self, path: Path) -> Book:
        return BookOpenpyxl(openpyxl.load_workbook(str(path.resolve())))
